from ..repositories.employee_repo import EmployeeRepository
from ..models import Employee
from bson.decimal128 import Decimal128
from ..serializers import EmployeeSerializer
import openpyxl
from django.db import transaction
from openpyxl import Workbook
from django.http import HttpResponse


class EmployeeService:
    @staticmethod
    def create_employee(data: dict) -> Employee:
        """
        Business logic for adding an employee.
        """
        print("we in services", data)
        mapping = {
            "firstName": "first_name",
            "lastName": "last_name",
            "department": "department",
            "position": "position",
            # "hireDate": "hire_date",
            "status": "status",
            "email": "email",
            "phone": "phone",
            "salary": "salary",
        }
        model_data = {mapping[k]: v for k, v in data.items() if k in mapping}
        existing_employee = EmployeeRepository.get_employee_by_email(
            model_data.get("email")
        )
        if existing_employee:
            raise ValueError("Employee with this email already exists.")

        employee = EmployeeRepository.create_employee(model_data)
        print("employee---- ", employee)
        return employee
        # serializer = EmployeeSerializer(employee)
        # return serializer.data

    @staticmethod
    def update_employee(id: int, data: dict) -> Employee:
        """
        Update employee
        """
        mapping = {
            "firstName": "first_name",
            "lastName": "last_name",
            "department": "department",
            "position": "position",
            # "hireDate": "hire_date",  # optional
            "status": "status",
            "email": "email",
            "phone": "phone",
            "salary": "salary",
        }
        model_data = {mapping[k]: v for k, v in data.items() if k in mapping}
        employee = EmployeeRepository.get_employee_by_id(id)
        if not employee:
            raise ValueError(f"Employee with id {id} does not exist.")

        # Optional: check for email conflict if email is being updated
        if "email" in model_data and model_data["email"] != employee.email:
            existing = EmployeeRepository.get_employee_by_email(model_data["email"])
            if existing:
                raise ValueError("Another employee with this email already exists.")

        # Update fields
        for key, value in model_data.items():
            setattr(employee, key, value)

        EmployeeRepository.save_employee(employee)
        return employee

    @staticmethod
    def get_employees() -> list[dict]:
        """
        Business logic for listing all existing employees.
        """
        employees = EmployeeRepository.get_all_employees()
        employee_list = []
        for employee in employees:
            salary = None
            if employee.salary is not None:
                if isinstance(employee.salary, Decimal128):
                    salary = float(employee.salary.to_decimal())
                else:
                    salary = float(employee.salary)
            employee_list.append(
                {
                    "id": employee.id,
                    "firstName": employee.first_name,
                    "lastName": employee.last_name,
                    "email": employee.email,
                    "phone": employee.phone,
                    "department": employee.department,
                    "position": employee.position,
                    # "hire_date": employee.hire_date.isoformat(),
                    "salary": salary,
                    "status": employee.status,
                }
            )

        return employee_list

    @staticmethod
    def delete_employee(employee_id):
        """
        Business logic for deleting employee by id,
        """
        employee = EmployeeRepository.get_employee_by_id(employee_id)

        if not employee:
            raise ValueError("Employee not found")

        return EmployeeRepository.delete_employee(employee)

    @staticmethod
    @transaction.atomic
    def import_from_excel(file):
        """
        Business logic for adding employees from imported excel to database.
        """
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        # Assumes row 1 is headers
        created = 0
        skipped = 0
        print("we here in import from excel")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                first_name,
                last_name,
                email,
                phone,
                department,
                position,
                hire_date,
                salary,
                status,
            ) = row

            # if EmployeeRepository.exists_by_email(email):
            #     skipped += 1
            #     continue

            # if hire_date and isinstance(hire_date, datetime.datetime):
            #     hire_date = hire_date.date()

            EmployeeRepository.create_employee(
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "phone": phone,
                    "department": department,
                    "position": position,
                    # "hire_date": hire_date,
                    "salary": salary,
                    "status": status,
                }
            )

            created += 1

        return {"created": created, "skipped": skipped}

    @staticmethod
    def export_to_excel():
        """
        Business logic for exporting database as excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "first_name",
            "last_name",
            "email",
            "phone",
            "department",
            "position",
            "hire_date",
            "salary",
            "status",
        ]
        ws.append(headers)

        employees = EmployeeRepository.get_all_employees()

        for emp in employees:
            salary_value = (
                float(emp.salary.to_decimal())
                if isinstance(emp.salary, Decimal128)
                else emp.salary
            )
            ws.append(
                [
                    emp.first_name,
                    emp.last_name,
                    emp.email,
                    emp.phone,
                    emp.department,
                    emp.position,
                    emp.hire_date,
                    # emp.salary,
                    salary_value,
                    emp.status,
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="employees.xlsx"'

        wb.save(response)
        return response
